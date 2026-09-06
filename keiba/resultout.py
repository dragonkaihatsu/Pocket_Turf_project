"""確定した結果を、読む用のTXTと表計算用のブックに書き出す。

予想と結果は `keiba.site.build_payload` が1つにまとめている。ここはその
出力を並べ替えるだけで、判定のやり直しはしない（同じ入力から同じ答えが
出るようにするため）。

    python3 -m keiba.cli results config/20260905_中央.json --output out/kekka

.xlsx は openpyxl があるときだけ。無ければ **BOM付きUTF-8のCSV** に落とす。
BOMが無いとWindowsのExcelがShift_JISと誤認して日本語が化ける。
"""
from __future__ import annotations

import csv
from pathlib import Path

STAKE = 100  # 1点あたりの賭け金

RACE_COLUMNS = ["日付", "場", "R", "レース名", "コース", "頭数",
                "1着馬番", "1着馬名", "1着人気", "1着印", "1着スコア順位",
                "2着馬番", "2着馬名", "2着印", "3着馬番", "3着馬名", "3着印",
                "買い目", "点数", "判定", "的中組", "配当", "投資", "払戻"]
CHAKU_COLUMNS = ["日付", "場", "R", "着順", "馬番", "馬名", "人気", "印", "スコア順位"]


def _hit_rows(race: dict) -> tuple[str, str, int | None]:
    """(判定, 的中組, 配当) を返す。"""
    b = (race.get("result") or {}).get("buy")
    if not b:
        return ("—", "", None)
    if b.get("hits"):
        h = b["hits"][0]
        return ("的中", h["combo"], h.get("pay"))
    return ("不的中", "", None)


def race_rows(payload: dict) -> list[dict]:
    """1行1レース。確定していないレースは出さない。"""
    date = payload.get("date") or ""
    out = []
    for r in payload["races"]:
        res = r.get("result")
        if not res:
            continue
        top = {t["chaku"]: t for t in res["top"]}
        rec = next((b for b in r.get("boxes", []) if b.get("rec")), None)
        judge, combo, pay = _hit_rows(r)
        points = rec["points"] if rec else 0
        row = {
            "日付": date, "場": r["venue"], "R": r["no"], "レース名": r["name"],
            "コース": r.get("surface", ""), "頭数": len(r["horses"]),
            "買い目": (f'{rec["kind"]} {rec["width"]}頭BOX' if rec else "—"),
            "点数": points, "判定": judge, "的中組": combo, "配当": pay,
            "投資": points * STAKE, "払戻": pay or 0,
        }
        for n in (1, 2, 3):
            t = top.get(n)
            row[f"{n}着馬番"] = t["umaban"] if t else ""
            row[f"{n}着馬名"] = t["name"] if t else ""
            row[f"{n}着印"] = (t.get("mark") or "") if t else ""
            if n == 1 and t:
                row["1着人気"] = t.get("ninki") or ""
                row["1着スコア順位"] = t.get("rank") or ""
        out.append({c: row.get(c, "") for c in RACE_COLUMNS})
    return out


def chaku_rows(payload: dict) -> list[dict]:
    """1行1着（1〜3着）。馬ごとに追う用。"""
    date = payload.get("date") or ""
    out = []
    for r in payload["races"]:
        res = r.get("result")
        if not res:
            continue
        for t in res["top"]:
            out.append({"日付": date, "場": r["venue"], "R": r["no"],
                        "着順": t["chaku"], "馬番": t["umaban"], "馬名": t["name"],
                        "人気": t.get("ninki") or "", "印": t.get("mark") or "",
                        "スコア順位": t.get("rank") or ""})
    return out


def summarize(rows: list[dict]) -> dict:
    inv = sum(r["投資"] or 0 for r in rows)
    ret = sum(r["払戻"] or 0 for r in rows)
    hit = sum(1 for r in rows if r["判定"] == "的中")
    return {"レース数": len(rows), "的中": hit,
            "的中率": (hit / len(rows)) if rows else 0.0,
            "投資": inv, "払戻": ret, "収支": ret - inv}


def format_text(payload: dict) -> str:
    """読む用。桁揃えはするが、ここはターミナル/メモ帳で見る前提。"""
    rows = race_rows(payload)
    L = [f'{payload.get("date", "")} 結果', ""]
    if not rows:
        L.append("確定したレースがありません")
        return "\n".join(L) + "\n"
    for r in rows:
        L.append(f'■ {r["場"]}{r["R"]} {r["レース名"]}　{r["コース"]}　{r["頭数"]}頭')
        for n in (1, 2, 3):
            mark = r[f"{n}着印"] or "－"
            extra = ""
            if n == 1:
                extra = f'　{r["1着人気"]}人気・スコア{r["1着スコア順位"]}位'
            L.append(f'  {n}着 {mark} {r[f"{n}着馬番"]} {r[f"{n}着馬名"]}{extra}')
        if r["判定"] == "的中":
            L.append(f'  {r["買い目"]} {r["点数"]}点 → 的中 {r["的中組"]}　'
                     f'{r["配当"]:,}円')
        else:
            L.append(f'  {r["買い目"]} {r["点数"]}点 → {r["判定"]}')
        L.append("")
    s = summarize(rows)
    L.append("── 合計 " + "─" * 30)
    L.append(f'{s["レース数"]}レース中 {s["的中"]}レース的中（{s["的中率"]:.0%}）')
    L.append(f'投資 {s["投資"]:,}円 / 払戻 {s["払戻"]:,}円 / 収支 {s["収支"]:+,}円')
    L.append("")
    L.append("※ 1点100円で計算。実際の投票額とは異なる場合があります")
    return "\n".join(L) + "\n"


def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    # BOM付きUTF-8。BOMが無いとWindowsのExcelがShift_JISと誤認して化ける
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        w.writerows(rows)


def write_book(path: Path, payload: dict) -> tuple[Path, bool]:
    """表計算ブックを書く。openpyxl が無ければ CSV に落として (path, False)。"""
    races = race_rows(payload)
    chaku = chaku_rows(payload)
    total = summarize(races)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        base = path.with_suffix("")
        write_csv(base.with_name(base.name + "_レース別.csv"), RACE_COLUMNS, races)
        write_csv(base.with_name(base.name + "_着順.csv"), CHAKU_COLUMNS, chaku)
        return (base.with_name(base.name + "_レース別.csv"), False)

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1B4B86")

    def sheet(title, columns, rows, widths):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        for r in rows:
            ws.append([r.get(c, "") for c in columns])
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.freeze_panes = "A2"
        return ws

    ws = wb.active
    ws.title = "サマリー"
    ws.append(["項目", "値"])
    for c in ws[1]:
        c.font = head_font
        c.fill = head_fill
    for k in ("レース数", "的中", "投資", "払戻", "収支"):
        ws.append([k, total[k]])
    ws.append(["的中率", round(total["的中率"], 4)])
    ws.append(["日付", payload.get("date", "")])
    ws.append(["注記", "1点100円で計算。実際の投票額とは異なる場合があります"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52

    sheet("レース別", RACE_COLUMNS, races,
          [11, 7, 6, 22, 18, 6] + [8, 16, 7, 5, 10] + [8, 16, 5] * 2
          + [15, 6, 7, 10, 9, 9, 9])
    sheet("着順", CHAKU_COLUMNS, chaku, [11, 7, 6, 6, 6, 18, 6, 5, 9])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return (path, True)
